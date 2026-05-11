# app.py - Truckload Lead Time Optimization Tool
# Sibling of the Ocean LTO tool. All ranking + statistical logic identical.
# TL-specific deltas:
#   - 4 always-available milestones (A/B/C/D) regardless of stop count
#   - Destination = dynamically detected last stop (Stop 1..20)
#   - No P44 fallbacks, no MSID min/max rollup
#   - Whole Journey ON => each unique stop chain becomes its own lane
#   - Negative segment durations excluded per-metric only (shipment kept for other metrics)
#   - Excel + ZIP-of-CSVs export, with a Key/Glossary sheet
import io
import re
import zipfile
from typing import Dict, Optional, List, Tuple

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st


# ============================================================
# Config
# ============================================================
MAX_STOPS = 20
TARGET_MODE = "TRUCKLOAD"

# Four always-available milestone keys. Position in stop chain is dynamic.
#   A = Origin Arrival             (first stop arrival)
#   B = Origin Departure           (first stop departure)
#   C = Destination Arrival        (last  stop arrival)
#   D = Destination Departure      (last  stop departure)
MILESTONES = ["A", "B", "C", "D"]
MILESTONE_LABELS = {
    "A": "Origin Arrival",
    "B": "Origin Departure",
    "C": "Destination Arrival",
    "D": "Destination Departure",
}
MILESTONE_INDEX = {ms: i for i, ms in enumerate(MILESTONES)}

# Source column names in the p44 unified shipment export.
COL_SHIPMENT_ID = "Shipment ID"
COL_MODE = "Current mode"
COL_CARRIER_NAME = "Current carrier"
COL_BOL = "Bill of lading"

# Tenant column (optional - p44 export usually has a single-tenant view)
# If absent, we synthesize a placeholder tenant so downstream grouping logic works.
TENANT_PLACEHOLDER = "TL Tenant"


def stop_col_name(n: int) -> str:
    return f"Stop {n}"


def stop_col_city(n: int) -> str:
    return f"Stop {n} city"


def stop_col_state(n: int) -> str:
    return f"Stop {n} state"


def stop_col_country(n: int) -> str:
    return f"Stop {n} country"


def stop_col_arrival(n: int) -> str:
    return f"Stop {n} actual arrival time"


def stop_col_departure(n: int) -> str:
    return f"Stop {n} actual departure time"


# Display labels for output column rename (sheet headers in Excel/CSV)
DISPLAY_COLS = {
    "TENANT_NAME": "Tenant Name",
    "LANE": "Lane",
    "CARRIER_NAME": "Carrier Name",
    "CARRIER_SCAC": "Carrier SCAC",
    "VOLUME": "Volume (Shipments)",
    "TOTAL_H": "Total Lead Time (Hours)",
    "TOTAL_D": "Total Lead Time (Days)",
    "MIN_H": "Min Lead Time (Hours)",
    "MIN_D": "Min Lead Time (Days)",
    "MED_H": "Median Lead Time (Hours)",
    "MED_D": "Median Lead Time (Days)",
    "PCT_H": "P{p} Lead Time (Hours)",
    "PCT_D": "P{p} Lead Time (Days)",
    "MAX_H": "Max Lead Time (Hours)",
    "MAX_D": "Max Lead Time (Days)",
}

DEFAULT_PERCENTILE_VOLUME_THRESHOLD_PCT = 0.0
DEFAULT_RECOMMENDATION_VOLUME_THRESHOLD_PCT = 0.0


# ============================================================
# File ingest (robust for big CSV / Excel uploads)
# ============================================================
def _columns_to_load(all_cols: List[str]) -> List[str]:
    """Return the subset of useful columns to actually read from disk.
    Drops everything we don't need so 100MB+ exports load fast."""
    wanted = {
        COL_SHIPMENT_ID,
        COL_MODE,
        COL_CARRIER_NAME,
        COL_BOL,
    }
    for n in range(1, MAX_STOPS + 1):
        wanted.update({
            stop_col_name(n),
            stop_col_city(n),
            stop_col_state(n),
            stop_col_country(n),
            stop_col_arrival(n),
            stop_col_departure(n),
        })
    # Tenant column is optional - include if present under common names
    optional_tenant_names = {"Tenant", "Tenant Name", "TENANT_NAME"}
    return [c for c in all_cols if c in wanted or c in optional_tenant_names]


def _slim_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Drop stop columns where every row is blank/NaN. The p44 unified export
    always includes Stop 1..20 columns, but most truckload shipments only use
    2-3 stops. Trimming the unused stop columns slashes memory by 60-80%
    on real uploads.

    Also category-encodes low-cardinality string columns (carrier, mode,
    country, state) to reduce string-object overhead on big uploads."""
    if df is None or len(df) == 0:
        return df
    to_drop = []
    for n in range(1, MAX_STOPS + 1):
        for col_fn in (stop_col_name, stop_col_city, stop_col_state,
                       stop_col_country, stop_col_arrival, stop_col_departure):
            c = col_fn(n)
            if c not in df.columns:
                continue
            s = df[c]
            # Fast emptiness check: NaN or string blanks/single-space.
            if pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
                non_blank = s.notna() & (s.astype(str).str.strip() != "")
            else:
                non_blank = s.notna()
            if not bool(non_blank.any()):
                to_drop.append(c)
    if to_drop:
        df = df.drop(columns=to_drop)

    # Category-encode columns we know are low cardinality. Saves 50-80% of
    # the string-object overhead on 95k-row uploads. Carrier name is usually
    # under 200 unique values; mode/country/state are tiny. NOTE: we do NOT
    # categorize Stop N name (often near-unique) or any timestamp/ID column.
    cat_candidates = [COL_MODE, COL_CARRIER_NAME]
    for n in range(1, MAX_STOPS + 1):
        cat_candidates.append(stop_col_state(n))
        cat_candidates.append(stop_col_country(n))
    for c in cat_candidates:
        if c in df.columns and df[c].dtype == "object":
            try:
                # Only convert if cardinality is comfortably below row count.
                nunique = df[c].nunique(dropna=True)
                if nunique > 0 and nunique < max(1024, len(df) // 4):
                    df[c] = df[c].astype("category")
            except Exception:
                pass
    return df


def _filter_to_truckload(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only TRUCKLOAD rows. Done as early as possible to reduce memory
    before any per-row work."""
    if COL_MODE not in df.columns:
        return df
    mode_norm = df[COL_MODE].astype(str).str.strip().str.upper()
    return df[mode_norm == TARGET_MODE].copy()


@st.cache_data(show_spinner=False, max_entries=2)
def _read_input(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Read the uploaded file (CSV or Excel) into a DataFrame.
    Cached on the raw bytes + name so re-runs don't re-parse the file.

    Memory optimizations applied at load time:
    - Read only the subset of columns we use (usecols).
    - Filter to TRUCKLOAD rows immediately (the export contains all modes).
    - Drop stop columns that are entirely blank (most TL shipments are 2-stop).
    """
    name = file_name.lower()
    bio = io.BytesIO(file_bytes)

    if name.endswith(".csv"):
        encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
        last_err = None
        # First pass: read header only to compute use-cols
        for enc in encodings:
            try:
                bio.seek(0)
                header = pd.read_csv(bio, encoding=enc, nrows=0)
                usecols = _columns_to_load(list(header.columns))
                bio.seek(0)
                df = pd.read_csv(bio, encoding=enc, usecols=usecols, low_memory=False)
                df = _filter_to_truckload(df)
                df = _slim_dataframe(df)
                return df
            except UnicodeDecodeError as e:
                last_err = e
                continue
        raise ValueError(
            f"Unable to decode CSV with {encodings}. Last error: {last_err}. "
            f"Try exporting as UTF-8, or upload Excel instead."
        )

    if name.endswith(".xlsx") or name.endswith(".xls"):
        bio.seek(0)
        header = pd.read_excel(bio, nrows=0)
        usecols = _columns_to_load(list(header.columns))
        bio.seek(0)
        df = pd.read_excel(bio, usecols=usecols)
        df = _filter_to_truckload(df)
        df = _slim_dataframe(df)
        return df

    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


def _normalize_cell(val) -> Optional[str]:
    """Convert a cell to a stripped string, treating blanks/single-space as None."""
    if val is None:
        return None
    if isinstance(val, float) and np.isnan(val):
        return None
    s = str(val).strip()
    if not s:
        return None
    return s


# ============================================================
# Tenant + carrier hygiene
# ============================================================
def ensure_tenant_column(df: pd.DataFrame) -> pd.DataFrame:
    """The p44 TL export typically has no Tenant column.
    If a column resembling 'Tenant' exists, copy it; otherwise insert placeholder."""
    if "TENANT_NAME" in df.columns:
        return df
    for candidate in ["Tenant Name", "Tenant", "TENANT_NAME"]:
        if candidate in df.columns:
            df["TENANT_NAME"] = df[candidate].fillna(TENANT_PLACEHOLDER).astype(str)
            return df
    df["TENANT_NAME"] = TENANT_PLACEHOLDER
    return df


def ensure_carrier_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map p44 'Current carrier' -> CARRIER_NAME, default CARRIER_SCAC to blank
    (the export doesn't include a TL SCAC field)."""
    df["CARRIER_NAME"] = df[COL_CARRIER_NAME].astype(str).where(
        df[COL_CARRIER_NAME].notna(), other="UNKNOWN"
    )
    df["CARRIER_SCAC"] = ""
    return df


# ============================================================
# Dynamic last-stop detection
# ============================================================
def detect_last_stop_index(df: pd.DataFrame) -> pd.Series:
    """For each row, return the largest N in 1..MAX_STOPS where Stop N has a
    populated name (treating ' ' and '' as blank). Stop 1 is always considered
    populated for a valid TL shipment; if Stop 1 is blank, returns NaN."""
    last_idx = pd.Series(np.nan, index=df.index, dtype="float64")
    BLANK_TOKENS = {"", "nan", "none", "null", "na", "n/a"}
    for n in range(1, MAX_STOPS + 1):
        col = stop_col_name(n)
        if col not in df.columns:
            continue
        # First filter out true NaN, then string-normalize and exclude blank tokens.
        raw = df[col]
        s = raw.where(raw.notna(), other="").astype(str).str.strip()
        populated = ~s.str.lower().isin(BLANK_TOKENS)
        last_idx = last_idx.where(~populated, other=float(n))
    return last_idx


def first_stop_index() -> int:
    return 1


# ============================================================
# Datetime helpers
# ============================================================
def _coerce_datetimes(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """Coerce timestamp columns to naive datetime64.

    The p44 unified shipment export emits timestamps formatted as
    ``"2026-05-09 01:09:01 IST"`` with a literal three-letter timezone
    abbreviation appended. ``pd.to_datetime`` with ``utc=True`` either rejects
    the suffix (yielding NaT for every row, which produced the
    "no eligible shipments" failure seen in production) or interprets ``IST``
    via the deprecated tzlocal fallback, which silently depends on the
    deployment host's timezone.

    Since every timestamp in the export carries the same timezone, we strip the
    trailing TZ abbreviation (or numeric offset) and parse the remainder as a
    naive datetime. Time deltas between two naive timestamps from the same
    source are identical to deltas computed in any consistent timezone, so
    lead-time math is unaffected. This also handles cells that contain a
    single space or other whitespace-only placeholders.
    """
    if df is None or len(df) == 0:
        return df
    # Regex strips: trailing whitespace + 2-5 letters (e.g. IST, UTC, EST,
    # PDT, GMT) OR trailing whitespace + numeric offset (+0530, -08:00).
    _TZ_SUFFIX_RE = r"\s+(?:[A-Za-z]{2,5}|[+-]\d{2}:?\d{2})\s*$"
    for c in cols:
        if c not in df.columns:
            continue
        col = df[c]
        # If already a datetime dtype, nothing to do.
        if pd.api.types.is_datetime64_any_dtype(col):
            # Drop any timezone info so downstream subtractions stay consistent.
            if getattr(col.dtype, "tz", None) is not None:
                df[c] = col.dt.tz_convert(None)
            continue
        # Coerce to string, treat NaN/None as empty, strip whitespace, drop
        # the literal " IST"-style suffix, then parse as naive.
        s = col.where(col.notna(), other="").astype(str).str.strip()
        s = s.str.replace(_TZ_SUFFIX_RE, "", regex=True)
        s = s.replace({"": None})
        # Fast path: p44 exports use "YYYY-MM-DD HH:MM:SS" after the TZ strip.
        # Try the explicit format first to avoid per-element dateutil fallback
        # (which is very slow on 100k+ row uploads). If anything fails the
        # fast path, fall back to the inferring parser for that column.
        parsed = pd.to_datetime(s, format="%Y-%m-%d %H:%M:%S", errors="coerce")
        nonnull_input = s.notna().sum()
        nonnull_parsed = parsed.notna().sum()
        if nonnull_input > 0 and nonnull_parsed < nonnull_input * 0.5:
            # Less than half parsed under the fast format: fall back to the
            # inferring parser so we still handle exotic formats correctly.
            parsed = pd.to_datetime(s, errors="coerce")
        df[c] = parsed
    return df


def _round_hours(x: Optional[float]) -> Optional[float]:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    return float(np.round(x, 2))


def _round_days_from_hours(x_hours: Optional[float]) -> Optional[int]:
    if x_hours is None or (isinstance(x_hours, float) and np.isnan(x_hours)):
        return None
    return int(np.round(x_hours / 24.0))


def _safe_quantile(values: pd.Series, q: float) -> Optional[float]:
    values = values.dropna()
    if values.empty:
        return None
    return float(values.quantile(q, interpolation="linear"))


def _pct_to_count(total_shipments: int, threshold_pct: float) -> int:
    if total_shipments <= 0 or threshold_pct <= 0:
        return 0
    return int(np.ceil(total_shipments * (threshold_pct / 100.0)))


# ============================================================
# Lane string builders
# ============================================================
def _format_endpoint(city, state, country) -> str:
    """Title-case city, upper-case state, upper-case country. Drop blanks cleanly.
    Returns '' if all three are blank."""
    parts = []
    if city is not None:
        c = str(city).strip()
        if c:
            parts.append(c.title())
    if state is not None:
        s = str(state).strip()
        if s:
            parts.append(s.upper())
    if country is not None:
        co = str(country).strip()
        if co:
            parts.append(co.upper())
    return ", ".join(parts)


def build_endpoint_for_row(row: pd.Series, stop_n: int) -> str:
    return _format_endpoint(
        row.get(stop_col_city(stop_n)),
        row.get(stop_col_state(stop_n)),
        row.get(stop_col_country(stop_n)),
    )


def build_chain_lane(row: pd.Series, last_idx: int) -> str:
    """Whole Journey ON: chain together every populated stop's endpoint."""
    parts = []
    for n in range(1, int(last_idx) + 1):
        ep = build_endpoint_for_row(row, n)
        if ep:
            parts.append(ep)
    return " → ".join(parts)


def build_direct_lane(row: pd.Series, last_idx: int) -> str:
    """Whole Journey OFF: first stop -> last stop only."""
    first_ep = build_endpoint_for_row(row, first_stop_index())
    last_ep = build_endpoint_for_row(row, int(last_idx))
    if not first_ep or not last_ep:
        return ""  # caller will drop the row
    return f"{first_ep} → {last_ep}"


# ---- Vectorized endpoint + lane helpers (hot path) -----------------------
def _endpoint_series_for_stop(df: pd.DataFrame, stop_n: int) -> pd.Series:
    """Vectorized: return a Series of formatted endpoint strings (one per row)
    for the given stop number. Empty when all three of city/state/country are
    blank for that row. Operates on string-vector primitives only - no apply."""
    city_col = stop_col_city(stop_n)
    state_col = stop_col_state(stop_n)
    country_col = stop_col_country(stop_n)

    def _norm(col_name: str, transform: str) -> pd.Series:
        if col_name not in df.columns:
            return pd.Series([""] * len(df), index=df.index, dtype="object")
        s = df[col_name]
        # Materialize categorical -> plain object strings before fill/replace,
        # since filling a Categorical with "" requires "" to be a category.
        if isinstance(s.dtype, pd.CategoricalDtype):
            s = s.astype(object)
        s = s.where(s.notna(), other="").astype(str).str.strip()
        if transform == "title":
            s = s.str.title()
        elif transform == "upper":
            s = s.str.upper()
        return s

    city = _norm(city_col, "title")
    state = _norm(state_col, "upper")
    country = _norm(country_col, "upper")

    has_city = city.str.len() > 0
    has_state = state.str.len() > 0
    has_country = country.str.len() > 0

    # Build comma-separated string with only the populated parts.
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    # Use a small fixed combinations loop (8 possibilities) for clarity + speed.
    for (c_has, s_has, co_has) in [
        (True, True, True),
        (True, True, False),
        (True, False, True),
        (True, False, False),
        (False, True, True),
        (False, True, False),
        (False, False, True),
        (False, False, False),
    ]:
        mask = (has_city == c_has) & (has_state == s_has) & (has_country == co_has)
        if not mask.any():
            continue
        if not (c_has or s_has or co_has):
            # Empty endpoint - already "" by default
            continue
        parts_cols = []
        if c_has:
            parts_cols.append(city[mask])
        if s_has:
            parts_cols.append(state[mask])
        if co_has:
            parts_cols.append(country[mask])
        # Series.str.cat would require alignment; build via concat + agg
        joined = parts_cols[0]
        for extra in parts_cols[1:]:
            joined = joined.str.cat(extra, sep=", ")
        out.loc[mask] = joined
    return out


def vectorized_direct_lanes(df: pd.DataFrame, last_idx: pd.Series) -> pd.Series:
    """Vectorized POL→POD lane string. Returns '' for rows whose first or
    last endpoint is blank (caller drops these)."""
    first_eps = _endpoint_series_for_stop(df, 1)
    last_eps = _gather_by_stop(df, last_idx, _endpoint_series_for_stop)
    has_both = (first_eps.str.len() > 0) & (last_eps.str.len() > 0)
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    out.loc[has_both] = first_eps[has_both].str.cat(last_eps[has_both], sep=" → ")
    return out


def _gather_by_stop(df: pd.DataFrame, last_idx: pd.Series, per_stop_series_fn) -> pd.Series:
    """Build a Series by selecting, for each row, the value of
    per_stop_series_fn(df, N) where N == last_idx[row]. Works without per-row
    apply by iterating stop numbers and copying into the output via mask."""
    out = None
    li = last_idx.astype("Int64") if not pd.api.types.is_integer_dtype(last_idx) else last_idx
    unique_idx = sorted({int(v) for v in li.dropna().unique()})
    for n in unique_idx:
        if n < 1 or n > MAX_STOPS:
            continue
        s_for_n = per_stop_series_fn(df, n)
        if out is None:
            # Default dtype matches the per-stop series dtype.
            out = pd.Series(index=df.index, dtype=s_for_n.dtype)
            if pd.api.types.is_object_dtype(s_for_n.dtype):
                out = out.where(out.notna(), other="")
        mask = (li == n).fillna(False)
        out.loc[mask] = s_for_n.loc[mask]
    if out is None:
        out = pd.Series([""] * len(df), index=df.index, dtype="object")
    return out


def _timestamp_series_for_stop(getter):
    """Returns a function(df, n) -> Series suitable for _gather_by_stop, where
    getter(n) yields the column name."""
    def _inner(df: pd.DataFrame, n: int) -> pd.Series:
        col = getter(n)
        if col in df.columns:
            return df[col]
        return pd.Series([pd.NaT] * len(df), index=df.index, dtype="datetime64[ns]")
    return _inner


def vectorized_milestone_ts(df: pd.DataFrame, last_idx: pd.Series, milestone: str) -> pd.Series:
    """Vectorized milestone timestamp resolver. A/B always at Stop 1, C/D at
    the dynamic last stop."""
    if milestone == "A":
        col = stop_col_arrival(1)
        if col in df.columns:
            return df[col].copy()
        return pd.Series([pd.NaT] * len(df), index=df.index, dtype="datetime64[ns]")
    if milestone == "B":
        col = stop_col_departure(1)
        if col in df.columns:
            return df[col].copy()
        return pd.Series([pd.NaT] * len(df), index=df.index, dtype="datetime64[ns]")
    if milestone == "C":
        return _gather_by_stop(df, last_idx, _timestamp_series_for_stop(stop_col_arrival))
    if milestone == "D":
        return _gather_by_stop(df, last_idx, _timestamp_series_for_stop(stop_col_departure))
    return pd.Series([pd.NaT] * len(df), index=df.index, dtype="datetime64[ns]")


def vectorized_chain_lanes(df: pd.DataFrame, last_idx: pd.Series) -> pd.Series:
    """Whole Journey ON chain lane: 'A → B → C → D' for each row's stops.
    Pre-computes endpoint strings per stop number once, then iterates over
    distinct last-stop counts in last_idx and joins per group."""
    li = last_idx.astype("Int64") if not pd.api.types.is_integer_dtype(last_idx) else last_idx
    max_n = int(li.max()) if pd.notna(li.max()) else 0
    # Pre-compute the endpoint Series for each stop position once.
    endpoint_per_n = {}
    for n in range(1, max_n + 1):
        endpoint_per_n[n] = _endpoint_series_for_stop(df, n)
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    for stop_count in sorted({int(v) for v in li.dropna().unique()}):
        if stop_count < 1:
            continue
        mask = (li == stop_count).fillna(False)
        if not mask.any():
            continue
        # For these rows, concatenate ep_1..ep_stop_count with " → " skipping blanks.
        # Approach: stack endpoint Series into a DataFrame slice, then row-wise
        # join via list comprehension on (already-vectorized) string columns.
        # This list comp loops only over rows in this group, not all rows.
        cols = [endpoint_per_n[n].loc[mask].values for n in range(1, stop_count + 1)]
        if not cols:
            continue
        # Transpose into (group_n_rows, stop_count) and join.
        # Using a Python loop on the small group is still much cheaper than
        # apply over the whole frame.
        joined = [
            " → ".join([str(v) for v in row_vals if v])
            for row_vals in zip(*cols)
        ]
        out.loc[mask] = joined
    return out


# ============================================================
# Milestone timestamp resolution (depends on each row's last-stop index)
# ============================================================
def resolve_milestone_ts(row: pd.Series, milestone: str, last_idx: int) -> pd.Timestamp:
    """Return the datetime for a given milestone (A/B/C/D) using the row's
    dynamically detected last stop index."""
    if milestone == "A":
        col = stop_col_arrival(first_stop_index())
    elif milestone == "B":
        col = stop_col_departure(first_stop_index())
    elif milestone == "C":
        col = stop_col_arrival(int(last_idx))
    elif milestone == "D":
        col = stop_col_departure(int(last_idx))
    else:
        return pd.NaT
    val = row.get(col, pd.NaT)
    return val if pd.notna(val) else pd.NaT


def required_milestone_columns(start_ms: str, end_ms: str, whole_journey: bool) -> List[str]:
    """List the column patterns required given the user's journey selection.
    For default journey only the chosen start/end milestone timestamps matter.
    For Whole Journey ON every stop arrival + departure matters (eligibility is
    checked at the row level since stop count is dynamic)."""
    required = set()
    # Always need first stop name + last stop name to define lane
    required.add(stop_col_name(1))
    if whole_journey:
        for n in range(1, MAX_STOPS + 1):
            required.add(stop_col_arrival(n))
            required.add(stop_col_departure(n))
    else:
        # Default journey: the picked start + end milestone timestamps
        for ms in (start_ms, end_ms):
            if ms == "A":
                required.add(stop_col_arrival(1))
            elif ms == "B":
                required.add(stop_col_departure(1))
            # C/D depend on the dynamic last-stop, can't be statically named here.
    return sorted(required)


def validate_input_columns(df: pd.DataFrame) -> List[str]:
    """Return list of missing critical columns. Returns [] if the file is OK."""
    missing = []
    must_have = [
        COL_SHIPMENT_ID,
        COL_MODE,
        COL_CARRIER_NAME,
        stop_col_name(1),
        stop_col_city(1),
        stop_col_state(1),
        stop_col_country(1),
        stop_col_arrival(1),
        stop_col_departure(1),
    ]
    for c in must_have:
        if c not in df.columns:
            missing.append(c)
    return missing


# ============================================================
# Core shipment computation
# ============================================================
def _segment_label_for(prev_n: int, next_n: int, last_idx: int) -> str:
    """Generate a human-readable segment label (transit between adjacent stops)."""
    if prev_n == 1:
        prev_name = "Origin"
    elif prev_n == int(last_idx):
        prev_name = "Destination"
    else:
        prev_name = f"Stop {prev_n}"
    if next_n == 1:
        next_name = "Origin"
    elif next_n == int(last_idx):
        next_name = "Destination"
    else:
        next_name = f"Stop {next_n}"
    return f"{prev_name} → {next_name}"


def _dwell_label_for(stop_n: int, last_idx: int) -> str:
    if stop_n == 1:
        return "Origin Dwell"
    if stop_n == int(last_idx):
        return "Destination Dwell"
    return f"Stop {stop_n} Dwell"


def _coerce_all_stop_timestamps(df: pd.DataFrame) -> pd.DataFrame:
    """Convert every Stop N arrival/departure column to datetime once, in place."""
    cols = []
    for n in range(1, MAX_STOPS + 1):
        if stop_col_arrival(n) in df.columns:
            cols.append(stop_col_arrival(n))
        if stop_col_departure(n) in df.columns:
            cols.append(stop_col_departure(n))
    return _coerce_datetimes(df, cols)


def compute_shipment_leadtimes(
    raw: pd.DataFrame,
    start_ms: str,
    end_ms: str,
    whole_journey: bool,
    progress_cb=None,
) -> pd.DataFrame:
    """Build a one-row-per-shipment frame with LANE + journey hours +
    (if whole_journey) every dwell/transit/total segment in hours.

    Negative durations are converted to NaN (per-metric exclusion) so the
    shipment is still counted in segments where its data is valid."""
    df = raw.copy()
    df = _coerce_all_stop_timestamps(df)
    df = ensure_tenant_column(df)
    df = ensure_carrier_columns(df)

    if progress_cb:
        progress_cb("Detecting last stop per shipment...", 0.10)
    df["_LAST_IDX"] = detect_last_stop_index(df)
    # Drop rows where Stop 1 is blank
    df = df[df["_LAST_IDX"].notna()].copy()
    df["_LAST_IDX"] = df["_LAST_IDX"].astype(int)

    if progress_cb:
        progress_cb("Building lane strings...", 0.20)

    # VECTORIZED lane construction. The previous version called
    # df.apply(axis=1) once per row, which on a 100k-row TL upload spawned a
    # Series per row and quickly exhausted Streamlit Cloud's 1 GB RAM.
    if whole_journey:
        df["LANE"] = vectorized_chain_lanes(df, df["_LAST_IDX"])
    else:
        df["LANE"] = vectorized_direct_lanes(df, df["_LAST_IDX"])

    # Drop shipments whose lane is empty (fully blank first or last endpoint)
    df = df[df["LANE"].astype(str).str.strip() != ""].copy()

    if df.empty:
        return pd.DataFrame(columns=[
            "TENANT_NAME", "MASTER_SHIPMENT_ID", "POL", "POD", "LANE",
            "CARRIER_NAME", "CARRIER_SCAC", "JOURNEY_LEAD_HOURS", "_LAST_IDX",
        ])

    if progress_cb:
        progress_cb("Resolving journey start/end timestamps...", 0.35)

    # VECTORIZED milestone timestamp resolution (was apply axis=1, now a
    # single vector op per milestone family).
    df["_START_TS"] = vectorized_milestone_ts(df, df["_LAST_IDX"], start_ms)
    df["_END_TS"] = vectorized_milestone_ts(df, df["_LAST_IDX"], end_ms)
    df["_START_TS"] = pd.to_datetime(df["_START_TS"], errors="coerce")
    df["_END_TS"] = pd.to_datetime(df["_END_TS"], errors="coerce")

    # Default journey eligibility - just need both selected milestones present
    rows_before = len(df)
    start_present = int(df["_START_TS"].notna().sum())
    end_present = int(df["_END_TS"].notna().sum())
    base_mask = df["_START_TS"].notna() & df["_END_TS"].notna()
    df = df[base_mask].copy()

    if df.empty:
        empty = pd.DataFrame(columns=[
            "TENANT_NAME", "MASTER_SHIPMENT_ID", "POL", "POD", "LANE",
            "CARRIER_NAME", "CARRIER_SCAC", "JOURNEY_LEAD_HOURS", "_LAST_IDX",
        ])
        empty.attrs["diagnostics"] = {
            "rows_with_lane": rows_before,
            "rows_with_start_ts": start_present,
            "rows_with_end_ts": end_present,
            "start_milestone": start_ms,
            "end_milestone": end_ms,
        }
        return empty

    if progress_cb:
        progress_cb("Computing journey lead times...", 0.50)

    journey_h = (df["_END_TS"] - df["_START_TS"]).dt.total_seconds() / 3600.0
    # Per-metric exclusion: any negative -> NaN
    journey_h = journey_h.where(journey_h >= 0, other=np.nan)
    df["JOURNEY_LEAD_HOURS"] = journey_h

    # POL/POD = first/last endpoint strings (vectorized).
    df["POL"] = _endpoint_series_for_stop(df, 1)
    df["POD"] = _gather_by_stop(df, df["_LAST_IDX"], _endpoint_series_for_stop)

    df["MASTER_SHIPMENT_ID"] = df[COL_SHIPMENT_ID].astype(str)

    if whole_journey:
        if progress_cb:
            progress_cb("Computing per-stop dwells and transits...", 0.65)
        df = _attach_segment_durations(df)

    keep_cols = [
        "TENANT_NAME", "MASTER_SHIPMENT_ID", "POL", "POD", "LANE",
        "CARRIER_NAME", "CARRIER_SCAC", "JOURNEY_LEAD_HOURS", "_LAST_IDX",
    ]
    segment_cols = [c for c in df.columns if c.startswith("SEG_") or c.startswith("DWELL_")]
    keep_cols = keep_cols + segment_cols
    keep_cols = [c for c in keep_cols if c in df.columns]
    # Slim down: a copy() materializes only the kept columns. Releases the
    # 200+ raw stop columns from memory once compute is done.
    result = df[keep_cols].copy()
    del df
    return result


def _attach_segment_durations(df: pd.DataFrame) -> pd.DataFrame:
    """In Whole Journey ON mode: for each shipment, compute every named
    segment (dwell at each stop + transit between adjacent stops) in hours.
    Negative durations -> NaN (per-metric exclusion)."""
    # Generate one column per possible segment for each stop count present in df.
    # Columns are keyed by an absolute segment ID (e.g. DWELL_1, DWELL_2,
    # SEG_1_2, SEG_2_3) - actual labels resolved later per lane.
    for n in range(1, MAX_STOPS + 1):
        a_col = stop_col_arrival(n)
        d_col = stop_col_departure(n)
        if a_col in df.columns and d_col in df.columns:
            durations = (df[d_col] - df[a_col]).dt.total_seconds() / 3600.0
            durations = durations.where(durations >= 0, other=np.nan)
            df[f"DWELL_{n}"] = durations
    for n in range(1, MAX_STOPS):
        d_col = stop_col_departure(n)
        a_next_col = stop_col_arrival(n + 1)
        if d_col in df.columns and a_next_col in df.columns:
            durations = (df[a_next_col] - df[d_col]).dt.total_seconds() / 3600.0
            durations = durations.where(durations >= 0, other=np.nan)
            df[f"SEG_{n}_{n + 1}"] = durations
    return df


# ============================================================
# Counts
# ============================================================
def compute_lane_and_carrier_counts(shipment_lt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if shipment_lt.empty:
        lane_counts = pd.DataFrame(columns=["Tenant Name", "Lane", "Shipments"])
        carrier_counts = pd.DataFrame(columns=["Tenant Name", "Carrier Name", "Carrier SCAC", "Shipments"])
        return lane_counts, carrier_counts

    lane_counts = (
        shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(columns={"TENANT_NAME": "Tenant Name", "LANE": "Lane", "MASTER_SHIPMENT_ID": "Shipments"})
        .sort_values(["Shipments", "Lane"], ascending=[False, True])
    )

    carrier_counts = (
        shipment_lt.groupby(["TENANT_NAME", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(
            columns={
                "TENANT_NAME": "Tenant Name",
                "CARRIER_NAME": "Carrier Name",
                "CARRIER_SCAC": "Carrier SCAC",
                "MASTER_SHIPMENT_ID": "Shipments",
            }
        )
        .sort_values(["Shipments", "Carrier Name"], ascending=[False, True])
    )
    return lane_counts, carrier_counts


def apply_top_n_lanes_filter(shipment_lt: pd.DataFrame, top_n_lanes: int) -> pd.DataFrame:
    if shipment_lt.empty or top_n_lanes <= 0:
        return shipment_lt

    lane_vol = (
        shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(columns={"MASTER_SHIPMENT_ID": "SHIPMENTS"})
    )
    top_lanes = (
        lane_vol.sort_values(["TENANT_NAME", "SHIPMENTS", "LANE"], ascending=[True, False, True])
        .groupby("TENANT_NAME", dropna=False)
        .head(top_n_lanes)[["TENANT_NAME", "LANE"]]
        .drop_duplicates()
    )
    return shipment_lt.merge(top_lanes, on=["TENANT_NAME", "LANE"], how="inner")


# ============================================================
# Report metrics (identical logic to ocean)
# ============================================================
def _stats_for_series(
    series: pd.Series,
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
    prefix: str,
) -> Dict[str, Optional[float]]:
    s = series.dropna()
    vol = int(s.shape[0])
    out = {
        f"{prefix}_TOTAL_H": None, f"{prefix}_TOTAL_D": None,
        f"{prefix}_MIN_H": None,   f"{prefix}_MIN_D": None,
        f"{prefix}_MED_H": None,   f"{prefix}_MED_D": None,
        f"{prefix}_PCT_H": None,   f"{prefix}_PCT_D": None,
        f"{prefix}_MAX_H": None,   f"{prefix}_MAX_D": None,
    }
    if vol == 0:
        return out
    total_h = _round_hours(float(s.sum()))
    min_h = _round_hours(float(s.min()))
    med_h = _round_hours(float(s.median()))
    max_h = _round_hours(float(s.max()))
    out[f"{prefix}_TOTAL_H"] = total_h
    out[f"{prefix}_TOTAL_D"] = _round_days_from_hours(total_h)
    out[f"{prefix}_MIN_H"] = min_h
    out[f"{prefix}_MIN_D"] = _round_days_from_hours(min_h)
    out[f"{prefix}_MED_H"] = med_h
    out[f"{prefix}_MED_D"] = _round_days_from_hours(med_h)
    out[f"{prefix}_MAX_H"] = max_h
    out[f"{prefix}_MAX_D"] = _round_days_from_hours(max_h)
    if include_percentile and vol >= int(min_volume_for_percentile):
        pct_h = _round_hours(_safe_quantile(s, percentile_p / 100.0))
        out[f"{prefix}_PCT_H"] = pct_h
        out[f"{prefix}_PCT_D"] = _round_days_from_hours(pct_h)
    return out


def build_duration_configs(
    start_ms: str, end_ms: str, whole_journey: bool, shipment_lt: pd.DataFrame
) -> List[Dict[str, str]]:
    """One config per metric column: the journey (always), plus every dwell +
    transit + total when Whole Journey is ON. Segment labels are derived from
    the data so labels read 'Origin Dwell', 'Stop 2 Dwell', 'Destination Dwell',
    'Origin → Stop 2', etc."""
    configs = [{
        "data_col": "JOURNEY_LEAD_HOURS",
        "prefix": "JOURNEY",
        "label": f"{MILESTONE_LABELS[start_ms]} → {MILESTONE_LABELS[end_ms]}",
        "display_mode": "journey",
    }]
    if not whole_journey or shipment_lt.empty:
        return configs

    last_idx_max = int(shipment_lt["_LAST_IDX"].max())
    # Dwells
    for n in range(1, last_idx_max + 1):
        col = f"DWELL_{n}"
        if col in shipment_lt.columns:
            configs.append({
                "data_col": col,
                "prefix": col,
                "label": _dwell_label_for(n, last_idx_max),
                "display_mode": "segment",
            })
    # Transits
    for n in range(1, last_idx_max):
        col = f"SEG_{n}_{n + 1}"
        if col in shipment_lt.columns:
            configs.append({
                "data_col": col,
                "prefix": col,
                "label": _segment_label_for(n, n + 1, last_idx_max),
                "display_mode": "segment",
            })
    return configs


def _group_stats(
    g: pd.DataFrame,
    duration_configs: List[Dict[str, str]],
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
) -> pd.Series:
    result = {"VOLUME": int(g["MASTER_SHIPMENT_ID"].nunique())}
    for cfg in duration_configs:
        col = cfg["data_col"]
        if col not in g.columns:
            for suf in ["_TOTAL_H", "_TOTAL_D", "_MIN_H", "_MIN_D", "_MED_H", "_MED_D",
                        "_PCT_H", "_PCT_D", "_MAX_H", "_MAX_D"]:
                result[f"{cfg['prefix']}{suf}"] = None
            continue
        result.update(
            _stats_for_series(
                g[col],
                percentile_p=percentile_p,
                include_percentile=include_percentile,
                min_volume_for_percentile=min_volume_for_percentile,
                prefix=cfg["prefix"],
            )
        )
    return pd.Series(result)


def build_carrier_lane_report(
    shipment_lt: pd.DataFrame,
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
    duration_configs: List[Dict[str, str]],
) -> pd.DataFrame:
    base_cols = ["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC", "VOLUME",
                 "_IS_LANE_ROW", "_POL", "_POD"]
    metric_cols = []
    for cfg in duration_configs:
        pfx = cfg["prefix"]
        metric_cols.extend([
            f"{pfx}_TOTAL_H", f"{pfx}_TOTAL_D",
            f"{pfx}_MIN_H",   f"{pfx}_MIN_D",
            f"{pfx}_MED_H",   f"{pfx}_MED_D",
            f"{pfx}_PCT_H",   f"{pfx}_PCT_D",
            f"{pfx}_MAX_H",   f"{pfx}_MAX_D",
        ])
    cols = base_cols + metric_cols
    if shipment_lt.empty:
        return pd.DataFrame(columns=cols)

    lane_cols = ["TENANT_NAME", "POL", "POD", "LANE"]
    lane_stats = (
        shipment_lt.groupby(lane_cols, dropna=False)
        .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_percentile, min_volume_for_percentile))
        .reset_index()
    )
    lane_stats["CARRIER_NAME"] = "ALL CARRIERS"
    lane_stats["CARRIER_SCAC"] = ""

    carrier_cols = ["TENANT_NAME", "POL", "POD", "LANE", "CARRIER_NAME", "CARRIER_SCAC"]
    carrier_stats = (
        shipment_lt.groupby(carrier_cols, dropna=False)
        .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_percentile, min_volume_for_percentile))
        .reset_index()
    )

    lane_stats = lane_stats.sort_values(["TENANT_NAME", "VOLUME", "LANE"], ascending=[True, False, True])

    rows = []
    for _, lr in lane_stats.iterrows():
        tenant, lane, pol, pod = lr["TENANT_NAME"], lr["LANE"], lr["POL"], lr["POD"]
        lane_row = {
            "TENANT_NAME": tenant, "LANE": lane,
            "CARRIER_NAME": lr["CARRIER_NAME"], "CARRIER_SCAC": lr["CARRIER_SCAC"],
            "VOLUME": lr["VOLUME"], "_IS_LANE_ROW": True, "_POL": pol, "_POD": pod,
        }
        for mc in metric_cols:
            lane_row[mc] = lr.get(mc)
        rows.append(lane_row)

        csub = carrier_stats[
            (carrier_stats["TENANT_NAME"] == tenant)
            & (carrier_stats["POL"].astype(str) == str(pol))
            & (carrier_stats["POD"].astype(str) == str(pod))
            & (carrier_stats["LANE"].astype(str) == str(lane))
        ].sort_values(["VOLUME", "CARRIER_NAME"], ascending=[False, True])

        for _, cr in csub.iterrows():
            carrier_row = {
                "TENANT_NAME": tenant, "LANE": "",
                "CARRIER_NAME": cr["CARRIER_NAME"], "CARRIER_SCAC": cr["CARRIER_SCAC"],
                "VOLUME": cr["VOLUME"], "_IS_LANE_ROW": False, "_POL": pol, "_POD": pod,
            }
            for mc in metric_cols:
                carrier_row[mc] = cr.get(mc)
            rows.append(carrier_row)
    return pd.DataFrame(rows, columns=cols)


# ============================================================
# Insights (identical MAD-based ranking to ocean)
# ============================================================
def build_insight_options(duration_configs: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    return {cfg["label"]: cfg for cfg in duration_configs}


def compute_insights_for_metric(
    shipment_lt: pd.DataFrame,
    metric_cfg: Dict[str, str],
    percentile_p: int,
    percentile_threshold_enabled: bool,
    percentile_threshold_pct: float,
    rec_threshold_enabled: bool,
    rec_threshold_pct: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Per-lane top-N carrier ranking by MAD-from-lane-median. Tiebreakers:
    lowest PXX absolute deviation -> higher shipment volume -> alphabetical."""
    metric_col = metric_cfg["data_col"]

    empty_lane_cols = [
        "TENANT_NAME", "LANE", "LANE_SHIPMENTS", "LANE_MEDIAN_H", "LANE_MEDIAN_D",
        "LANE_PXX_H", "LANE_PXX_D", "CARRIER_COUNT",
        "PERCENTILE_MIN_SHIPMENTS_REQUIRED", "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED",
    ]
    empty_carrier_cols = [
        "TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC", "SHIPMENTS",
        "CARRIER_SHARE_PCT", "CARRIER_MEDIAN_H", "CARRIER_MEDIAN_D",
        "CARRIER_PXX_H", "CARRIER_PXX_D", "MEDIAN_ABS_DEV_H", "MEDIAN_ABS_DEV_D",
        "PXX_ABS_DEV_H", "PXX_ABS_DEV_D", "RANK_IN_LANE",
        "RECOMMENDATION_ELIGIBLE", "PERCENTILE_ELIGIBLE",
        "PERCENTILE_MIN_SHIPMENTS_REQUIRED", "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED",
    ]
    if shipment_lt.empty or metric_col not in shipment_lt.columns:
        return pd.DataFrame(columns=empty_lane_cols), pd.DataFrame(columns=empty_carrier_cols)

    valid = shipment_lt.dropna(subset=[metric_col]).copy()
    if valid.empty:
        return pd.DataFrame(columns=empty_lane_cols), pd.DataFrame(columns=empty_carrier_cols)

    lane_base = (
        valid.groupby(["TENANT_NAME", "LANE"], dropna=False)
        .agg(
            LANE_SHIPMENTS=("MASTER_SHIPMENT_ID", "nunique"),
            LANE_MEDIAN_H=(metric_col, "median"),
            CARRIER_COUNT=("CARRIER_NAME", "nunique"),
        )
        .reset_index()
    )
    lane_base["PERCENTILE_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(percentile_threshold_pct)) if percentile_threshold_enabled else 0
    )
    lane_base["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(rec_threshold_pct)) if rec_threshold_enabled else 0
    )

    lane_pxx_rows = []
    for _, row in lane_base.iterrows():
        sub = valid[(valid["TENANT_NAME"] == row["TENANT_NAME"]) & (valid["LANE"] == row["LANE"])][metric_col]
        if percentile_threshold_enabled and int(row["LANE_SHIPMENTS"]) < int(row["PERCENTILE_MIN_SHIPMENTS_REQUIRED"]):
            lane_pxx = None
        else:
            lane_pxx = _safe_quantile(sub, percentile_p / 100.0)
        lane_pxx_rows.append(lane_pxx)
    lane_base["LANE_PXX_H"] = lane_pxx_rows
    lane_base["LANE_MEDIAN_H"] = lane_base["LANE_MEDIAN_H"].apply(_round_hours)
    lane_base["LANE_MEDIAN_D"] = lane_base["LANE_MEDIAN_H"].apply(_round_days_from_hours)
    lane_base["LANE_PXX_H"] = lane_base["LANE_PXX_H"].apply(_round_hours)
    lane_base["LANE_PXX_D"] = lane_base["LANE_PXX_H"].apply(_round_days_from_hours)

    merged = valid.merge(
        lane_base[["TENANT_NAME", "LANE", "LANE_MEDIAN_H", "LANE_SHIPMENTS",
                   "PERCENTILE_MIN_SHIPMENTS_REQUIRED", "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"]],
        on=["TENANT_NAME", "LANE"], how="left",
    )
    merged["ABS_DEV_H"] = (merged[metric_col] - merged["LANE_MEDIAN_H"]).abs()

    carrier_rows = []
    grouped = merged.groupby(["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)
    for (tenant, lane, carrier_name, carrier_scac), g in grouped:
        shipments = int(g["MASTER_SHIPMENT_ID"].nunique())
        lane_shipments = int(g["LANE_SHIPMENTS"].iloc[0]) if not g.empty else 0
        carrier_share_pct = round((shipments / lane_shipments * 100.0), 2) if lane_shipments > 0 else None
        pct_min_shipments = int(g["PERCENTILE_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0
        rec_min_shipments = int(g["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0

        carrier_series = g[metric_col].dropna()
        dev_series = g["ABS_DEV_H"].dropna()
        carrier_median_h = _round_hours(float(carrier_series.median())) if not carrier_series.empty else None
        mad_h = _round_hours(float(dev_series.median())) if not dev_series.empty else None

        percentile_eligible = not percentile_threshold_enabled or shipments >= pct_min_shipments
        recommendation_eligible = not rec_threshold_enabled or shipments >= rec_min_shipments

        if percentile_eligible:
            carrier_pxx_h = _round_hours(_safe_quantile(carrier_series, percentile_p / 100.0)) if not carrier_series.empty else None
            dev_pxx_h = _round_hours(_safe_quantile(dev_series, percentile_p / 100.0)) if not dev_series.empty else None
        else:
            carrier_pxx_h = None
            dev_pxx_h = None

        carrier_rows.append({
            "TENANT_NAME": tenant, "LANE": lane,
            "CARRIER_NAME": carrier_name, "CARRIER_SCAC": carrier_scac,
            "SHIPMENTS": shipments, "CARRIER_SHARE_PCT": carrier_share_pct,
            "CARRIER_MEDIAN_H": carrier_median_h, "CARRIER_MEDIAN_D": _round_days_from_hours(carrier_median_h),
            "CARRIER_PXX_H": carrier_pxx_h, "CARRIER_PXX_D": _round_days_from_hours(carrier_pxx_h),
            "MEDIAN_ABS_DEV_H": mad_h, "MEDIAN_ABS_DEV_D": _round_days_from_hours(mad_h),
            "PXX_ABS_DEV_H": dev_pxx_h, "PXX_ABS_DEV_D": _round_days_from_hours(dev_pxx_h),
            "RECOMMENDATION_ELIGIBLE": recommendation_eligible,
            "PERCENTILE_ELIGIBLE": percentile_eligible,
            "PERCENTILE_MIN_SHIPMENTS_REQUIRED": pct_min_shipments,
            "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED": rec_min_shipments,
        })

    carrier_recs = pd.DataFrame(carrier_rows)
    if carrier_recs.empty:
        carrier_recs["RANK_IN_LANE"] = pd.Series(dtype="Int64")
        return lane_base, carrier_recs

    carrier_recs["PXX_ABS_DEV_SORT"] = carrier_recs["PXX_ABS_DEV_H"].fillna(np.inf)
    carrier_recs["MEDIAN_ABS_DEV_SORT"] = carrier_recs["MEDIAN_ABS_DEV_H"].fillna(np.inf)

    eligible_recs = carrier_recs[carrier_recs["RECOMMENDATION_ELIGIBLE"]].copy()
    eligible_recs = eligible_recs.sort_values(
        ["TENANT_NAME", "LANE", "MEDIAN_ABS_DEV_SORT", "PXX_ABS_DEV_SORT", "SHIPMENTS", "CARRIER_NAME"],
        ascending=[True, True, True, True, False, True],
    )
    eligible_recs["RANK_IN_LANE"] = eligible_recs.groupby(["TENANT_NAME", "LANE"], dropna=False).cumcount() + 1

    non_eligible = carrier_recs[~carrier_recs["RECOMMENDATION_ELIGIBLE"]].copy()
    non_eligible["RANK_IN_LANE"] = pd.NA

    out = pd.concat([eligible_recs, non_eligible], ignore_index=True)
    out = out.drop(columns=["PXX_ABS_DEV_SORT", "MEDIAN_ABS_DEV_SORT"], errors="ignore")

    lane_summary = lane_base.sort_values(["LANE_SHIPMENTS", "LANE"], ascending=[False, True])
    carrier_recs = out.sort_values(
        ["TENANT_NAME", "LANE", "RECOMMENDATION_ELIGIBLE", "RANK_IN_LANE", "SHIPMENTS", "CARRIER_NAME"],
        ascending=[True, True, False, True, False, True],
    )
    return lane_summary, carrier_recs


def make_lane_selector_labels(lane_summary: pd.DataFrame) -> Dict[str, Tuple[str, str]]:
    mapping = {}
    if lane_summary.empty:
        return mapping
    for _, row in lane_summary.iterrows():
        label = f"{row['TENANT_NAME']} | {row['LANE']} ({int(row['LANE_SHIPMENTS'])} shipments)"
        mapping[label] = (row["TENANT_NAME"], row["LANE"])
    return mapping


def get_selected_lane_outputs(
    shipment_lt: pd.DataFrame,
    lane_summary: pd.DataFrame,
    carrier_recs: pd.DataFrame,
    selected_tenant: str,
    selected_lane: str,
    metric_cfg: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    metric_col = metric_cfg["data_col"]
    lane_row = lane_summary[(lane_summary["TENANT_NAME"] == selected_tenant) & (lane_summary["LANE"] == selected_lane)].copy()
    lane_carriers = carrier_recs[(carrier_recs["TENANT_NAME"] == selected_tenant) & (carrier_recs["LANE"] == selected_lane)].copy()

    top5 = lane_carriers[lane_carriers["RECOMMENDATION_ELIGIBLE"]].copy()
    top5 = top5.sort_values(["RANK_IN_LANE", "SHIPMENTS", "CARRIER_NAME"], ascending=[True, False, True]).head(5)

    carriers_to_plot = top5["CARRIER_NAME"].tolist()
    ship_subset = shipment_lt[
        (shipment_lt["TENANT_NAME"] == selected_tenant)
        & (shipment_lt["LANE"] == selected_lane)
        & (shipment_lt["CARRIER_NAME"].isin(carriers_to_plot))
    ][["TENANT_NAME", "LANE", "MASTER_SHIPMENT_ID", "CARRIER_NAME", "CARRIER_SCAC", metric_col]].copy()
    ship_subset = ship_subset.rename(columns={metric_col: "LEAD_TIME_HOURS"})
    lane_median_h = lane_row["LANE_MEDIAN_H"].iloc[0] if not lane_row.empty else None
    if lane_median_h is not None and not ship_subset.empty:
        ship_subset["ABS_DEV_H"] = (ship_subset["LEAD_TIME_HOURS"] - lane_median_h).abs()
    else:
        ship_subset["ABS_DEV_H"] = np.nan
    return lane_row, top5, ship_subset


# ============================================================
# Key / Glossary sheet content
# ============================================================
def build_key_glossary(duration_configs: List[Dict[str, str]], percentile_p: int) -> pd.DataFrame:
    rows = [
        ("Tenant Name", "Logical owner of the shipments. Defaults to '" + TENANT_PLACEHOLDER + "' if the export has no Tenant column."),
        ("Lane", "Whole Journey OFF: '<Origin City, ST, Country> → <Destination City, ST, Country>' (first stop → last stop only). Whole Journey ON: full ordered chain of every populated stop joined by ' → '. Each unique chain is its own lane."),
        ("Carrier Name", "Mapped from the export's 'Current carrier' field."),
        ("Carrier SCAC", "Always blank for TL (the p44 unified shipment export has no TL SCAC field)."),
        ("Volume (Shipments)", "Count of unique Shipment IDs in the group."),
        ("Median Lead Time (Hours/Days)", "Median (50th percentile) of the lead time series. Hours rounded to 2 dp, Days = Hours/24 rounded to nearest integer."),
        (f"P{percentile_p} Lead Time (Hours/Days)", "Percentile of the lead time series. Default 80th. Configurable from the sidebar."),
        ("Min / Max Lead Time", "Smallest and largest observed lead time in the group."),
        ("Total Lead Time", "Sum of lead times in the group (used internally; mostly informational at lane/carrier level)."),
        ("Median Abs Deviation (Hours/Days)", "Median absolute deviation of the carrier's shipments from the LANE median. Primary ranking metric: lower is better (more predictable carrier)."),
        ("P{p} Abs Deviation".format(p=percentile_p), "Same idea but using the percentile-band deviation. Secondary ranking metric."),
        ("Rank in Lane", "Carrier rank within its lane. Sort = lowest MAD → lowest PXX abs deviation → higher volume → alphabetical."),
        ("Recommendation Eligible", "TRUE if the carrier meets the volume threshold for being ranked (always TRUE when the threshold checkbox is off)."),
        ("Percentile Eligible", "TRUE if the carrier meets the volume threshold for its percentile metric to be computed."),
        ("Lane Median / Lane PXX", "Lane-level median + percentile of the lead time metric across ALL carriers in the lane."),
        ("Origin Dwell", "Stop 1 actual departure minus Stop 1 actual arrival (hours)."),
        ("Destination Dwell", "Last stop actual departure minus last stop actual arrival (hours)."),
        ("Stop N Dwell", "Intermediate stop N actual departure minus actual arrival (hours)."),
        ("Origin → Stop 2 (and other transits)", "Time in transit between two adjacent stops: arrival at Stop n+1 minus departure from Stop n (hours)."),
        ("Total Lead Time (in Whole Journey)", "End-to-end duration: from the chosen start milestone to the chosen end milestone."),
        ("Negative durations", "If any segment computes to a negative value (corrupted timestamps), that one segment is excluded for that shipment but the shipment is still counted in every other segment."),
        ("Milestone A", "Origin Arrival (Stop 1 actual arrival)."),
        ("Milestone B", "Origin Departure (Stop 1 actual departure)."),
        ("Milestone C", "Destination Arrival (last populated stop's actual arrival)."),
        ("Milestone D", "Destination Departure (last populated stop's actual departure)."),
    ]
    # Append per-segment label rows from the duration configs
    for cfg in duration_configs:
        if cfg["display_mode"] == "segment":
            rows.append((cfg["label"], f"Computed from column '{cfg['data_col']}'."))
    return pd.DataFrame(rows, columns=["Field / Term", "Definition"])


# ============================================================
# Excel writers
# ============================================================
def write_insights_excel(
    lane_summary: pd.DataFrame,
    carrier_recs: pd.DataFrame,
    selected_lane_shipments: pd.DataFrame,
    key_df: pd.DataFrame,
) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        lane_summary.to_excel(writer, sheet_name="Lane Summary", index=False)
        carrier_recs.to_excel(writer, sheet_name="Carrier Recommendations", index=False)
        selected_lane_shipments.to_excel(writer, sheet_name="Selected Lane Shipments", index=False)
        key_df.to_excel(writer, sheet_name="Key", index=False)
        for ws in writer.book.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 28
    output.seek(0)
    return output.getvalue()


def build_export_rename_map(duration_configs: List[Dict[str, str]], percentile_p: int) -> Dict[str, str]:
    export_cols = {
        "TENANT_NAME": DISPLAY_COLS["TENANT_NAME"],
        "LANE": DISPLAY_COLS["LANE"],
        "CARRIER_NAME": DISPLAY_COLS["CARRIER_NAME"],
        "CARRIER_SCAC": DISPLAY_COLS["CARRIER_SCAC"],
        "VOLUME": DISPLAY_COLS["VOLUME"],
    }
    for cfg in duration_configs:
        pfx = cfg["prefix"]
        label = cfg["label"]
        if cfg["display_mode"] == "journey":
            export_cols[f"{pfx}_TOTAL_H"] = DISPLAY_COLS["TOTAL_H"]
            export_cols[f"{pfx}_TOTAL_D"] = DISPLAY_COLS["TOTAL_D"]
            export_cols[f"{pfx}_MIN_H"] = DISPLAY_COLS["MIN_H"]
            export_cols[f"{pfx}_MIN_D"] = DISPLAY_COLS["MIN_D"]
            export_cols[f"{pfx}_MED_H"] = DISPLAY_COLS["MED_H"]
            export_cols[f"{pfx}_MED_D"] = DISPLAY_COLS["MED_D"]
            export_cols[f"{pfx}_PCT_H"] = DISPLAY_COLS["PCT_H"].format(p=percentile_p)
            export_cols[f"{pfx}_PCT_D"] = DISPLAY_COLS["PCT_D"].format(p=percentile_p)
            export_cols[f"{pfx}_MAX_H"] = DISPLAY_COLS["MAX_H"]
            export_cols[f"{pfx}_MAX_D"] = DISPLAY_COLS["MAX_D"]
        else:
            export_cols[f"{pfx}_TOTAL_H"] = f"{label} Total (Hours)"
            export_cols[f"{pfx}_TOTAL_D"] = f"{label} Total (Days)"
            export_cols[f"{pfx}_MIN_H"] = f"{label} Min (Hours)"
            export_cols[f"{pfx}_MIN_D"] = f"{label} Min (Days)"
            export_cols[f"{pfx}_MED_H"] = f"{label} Median (Hours)"
            export_cols[f"{pfx}_MED_D"] = f"{label} Median (Days)"
            export_cols[f"{pfx}_PCT_H"] = f"{label} P{percentile_p} (Hours)"
            export_cols[f"{pfx}_PCT_D"] = f"{label} P{percentile_p} (Days)"
            export_cols[f"{pfx}_MAX_H"] = f"{label} Max (Hours)"
            export_cols[f"{pfx}_MAX_D"] = f"{label} Max (Days)"
    return export_cols


def write_excel_counts(lane_counts: pd.DataFrame, carrier_counts: pd.DataFrame, key_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        lane_counts.to_excel(writer, sheet_name="Lane Counts", index=False)
        carrier_counts.to_excel(writer, sheet_name="Carrier Counts", index=False)
        key_df.to_excel(writer, sheet_name="Key", index=False)
        bold = Font(bold=True)
        for sheet_name in ["Lane Counts", "Carrier Counts", "Key"]:
            ws = writer.book[sheet_name]
            for cell in ws[1]:
                cell.font = bold
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 28
    output.seek(0)
    return output.getvalue()


def write_excel_final(
    raw_df: pd.DataFrame,
    report_df: pd.DataFrame,
    duration_configs: List[Dict[str, str]],
    percentile_p: int,
    key_df: pd.DataFrame,
    include_raw_data: bool = False,
) -> bytes:
    """Build the final Carrier Lane Lead Excel report.

    The Raw Data sheet is OPTIONAL (``include_raw_data``). It used to be on
    by default, but on a 65k-row upload openpyxl can balloon to 500-800 MB
    in memory while writing the sheet, which is what was crashing the
    Streamlit Cloud worker. The Carrier Lane Lead + Key sheets together are
    only a few thousand rows and well under 10 MB."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if include_raw_data:
            raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
        export_rename_map = build_export_rename_map(duration_configs, percentile_p)

        if report_df.empty:
            pd.DataFrame(columns=list(export_rename_map.values())).to_excel(writer, sheet_name="Carrier Lane Lead", index=False)
        else:
            df = report_df.copy()
            lane_flags = df["_IS_LANE_ROW"].astype(bool).to_list()
            df = df.drop(columns=["_IS_LANE_ROW", "_POL", "_POD"], errors="ignore")
            ordered_export_keys = [k for k in export_rename_map.keys() if k in df.columns]
            df = df[ordered_export_keys].rename(columns=export_rename_map)
            df.to_excel(writer, sheet_name="Carrier Lane Lead", index=False)

            ws = writer.book["Carrier Lane Lead"]
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
            lane_col_idx = list(df.columns).index(DISPLAY_COLS["LANE"]) + 1
            for i, is_lane in enumerate(lane_flags, start=2):
                if is_lane:
                    ws.cell(row=i, column=lane_col_idx).font = bold_font
            for idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 26

        key_df.to_excel(writer, sheet_name="Key", index=False)
        if include_raw_data:
            ws_raw = writer.book["Raw Data"]
            for cell in ws_raw[1]:
                cell.font = Font(bold=True)
        ws_key = writer.book["Key"]
        for cell in ws_key[1]:
            cell.font = Font(bold=True)
        for idx in range(1, ws_key.max_column + 1):
            ws_key.column_dimensions[get_column_letter(idx)].width = 60 if idx == 2 else 32

    output.seek(0)
    return output.getvalue()


def write_csv_zip(*named_frames: Tuple[str, pd.DataFrame], key_df: Optional[pd.DataFrame] = None) -> bytes:
    """Pack multiple DataFrames + Key as one ZIP of CSVs."""
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, df in named_frames:
            csv_bytes = df.to_csv(index=False).encode("utf-8")
            zf.writestr(name, csv_bytes)
        if key_df is not None:
            zf.writestr("key.csv", key_df.to_csv(index=False).encode("utf-8"))
    output.seek(0)
    return output.getvalue()


# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(page_title="Truckload Lead Time Analyzer", layout="wide")
st.title("Truckload Lead Time Analyzer (Carrier + Lane Lead Time)")

st.markdown(
    """
Upload a **CSV or Excel** extract from the p44 unified shipment export.
Pick journey **start** and **end** milestones from the four always-available milestones
(Origin Arrival, Origin Departure, Destination Arrival, Destination Departure).

Output includes:
- **Raw Data**
- **Carrier Lane Lead**
- **Insights**
- **Key / Glossary**
"""
)

uploaded = st.file_uploader("Upload CSV or Excel", type=["csv", "xlsx", "xls"])
if uploaded is None:
    st.stop()

# Step 1: Read file
load_status = st.status("Reading file...", expanded=False)
load_progress = st.progress(0.0)
try:
    file_bytes = uploaded.getvalue()
    load_progress.progress(0.30)
    raw_df = _read_input(file_bytes, uploaded.name)
    load_progress.progress(1.0)
    load_status.update(label=f"Loaded {raw_df.shape[0]:,} rows × {raw_df.shape[1]:,} columns", state="complete")
except Exception as e:
    load_progress.empty()
    load_status.update(label="File read failed", state="error")
    st.error(f"Could not read file: {e}")
    st.stop()

# Step 2: Validate columns
missing_cols = validate_input_columns(raw_df)
if missing_cols:
    st.error(
        "The uploaded file is missing required columns:\n\n- " + "\n- ".join(missing_cols)
        + "\n\nThis tool expects the p44 unified shipment export format."
    )
    st.stop()

# Step 3: TRUCKLOAD filter is now applied inside _read_input() to lower peak
# memory at load time. We still surface a friendly empty-file message here.
if raw_df.empty:
    st.error("No TRUCKLOAD shipments found in the upload.")
    st.stop()

st.success(f"Working with {raw_df.shape[0]:,} TRUCKLOAD rows × {raw_df.shape[1]:,} columns")

# ----------- Sidebar -----------
st.sidebar.header("Journey Settings")

start_ms = st.sidebar.selectbox(
    "Journey start milestone",
    MILESTONES,
    index=MILESTONES.index("B"),
    format_func=lambda ms: f"{ms} - {MILESTONE_LABELS[ms]}",
)
end_ms = st.sidebar.selectbox(
    "Journey end milestone",
    MILESTONES,
    index=MILESTONES.index("C"),
    format_func=lambda ms: f"{ms} - {MILESTONE_LABELS[ms]}",
)

# Chronological order check
if MILESTONE_INDEX[end_ms] <= MILESTONE_INDEX[start_ms]:
    st.error(
        f"End milestone must come AFTER start milestone in the natural order "
        f"(A → B → C → D). You selected {start_ms} → {end_ms}."
    )
    st.stop()

whole_journey = st.sidebar.checkbox("Calculate for whole journey", value=False)

st.sidebar.divider()
st.sidebar.header("Lane Filter")
top_n_lanes = st.sidebar.number_input(
    "Limit analysis to Top N lanes by volume (0 = all lanes)",
    min_value=0, max_value=1000, value=0, step=5,
)

st.sidebar.divider()
st.sidebar.header("Insights Settings")
include_percentile = st.sidebar.checkbox("Include additional percentile (PXX)", value=True)
percentile_p = st.sidebar.number_input(
    "Percentile value (e.g., 80)",
    min_value=1, max_value=99, value=80, step=1,
    disabled=not include_percentile,
)
limit_by_volume = st.sidebar.checkbox(
    "Only compute percentile if carrier volume share ≥ threshold (%)",
    value=False, disabled=not include_percentile,
)
percentile_volume_threshold_pct = st.sidebar.number_input(
    "Percentile volume threshold (%)",
    min_value=0.0, max_value=100.0,
    value=float(DEFAULT_PERCENTILE_VOLUME_THRESHOLD_PCT), step=1.0,
    disabled=(not include_percentile) or (not limit_by_volume),
)
recommendation_threshold_enabled = st.sidebar.checkbox(
    "Only generate recommendations if carrier volume share ≥ threshold (%)",
    value=False,
)
recommendation_threshold_pct = st.sidebar.number_input(
    "Recommendation volume threshold (%)",
    min_value=0.0, max_value=100.0,
    value=float(DEFAULT_RECOMMENDATION_VOLUME_THRESHOLD_PCT), step=1.0,
    disabled=not recommendation_threshold_enabled,
)

# ----------- Compute (cached on file + journey settings) -----------
# Streamlit reruns the entire script on every widget interaction. Caching
# compute_shipment_leadtimes() keyed on (file_id, start, end, whole_journey)
# means changing downstream-only knobs (top N lanes, percentile, threshold
# checkboxes) does NOT trigger a recompute, which keeps the UI snappy and
# avoids unnecessary memory churn on big uploads.
_compute_cache_key = (
    uploaded.name, len(file_bytes),
    str(start_ms), str(end_ms), bool(whole_journey),
)
if (
    st.session_state.get("_compute_cache_key") == _compute_cache_key
    and st.session_state.get("_compute_cache_result") is not None
):
    shipment_lt_all = st.session_state["_compute_cache_result"]
    compute_status = st.status("Using cached lead times", expanded=False, state="complete")
    compute_progress = st.progress(1.0)
else:
    compute_status = st.status("Computing shipment-level lead times...", expanded=False)
    compute_progress = st.progress(0.0)

    def _progress(label: str, frac: float):
        compute_progress.progress(min(max(frac, 0.0), 1.0))
        compute_status.update(label=label)

    try:
        shipment_lt_all = compute_shipment_leadtimes(
            raw=raw_df,
            start_ms=start_ms,
            end_ms=end_ms,
            whole_journey=whole_journey,
            progress_cb=_progress,
        )
        compute_status.update(label="Lead times computed", state="complete")
        st.session_state["_compute_cache_key"] = _compute_cache_key
        st.session_state["_compute_cache_result"] = shipment_lt_all
    except Exception as e:
        compute_progress.empty()
        compute_status.update(label="Compute failed", state="error")
        st.error(f"Error computing lead times: {e}")
        st.stop()

# Top-N filter is cheap (just a groupby + isin) - always re-run on change.
shipment_lt = apply_top_n_lanes_filter(shipment_lt_all, int(top_n_lanes))

# ----------- Metrics row -----------
total_shipments_raw = raw_df[COL_SHIPMENT_ID].nunique() if COL_SHIPMENT_ID in raw_df.columns else None
eligible_shipments = shipment_lt["MASTER_SHIPMENT_ID"].nunique() if not shipment_lt.empty else 0
coverage = (eligible_shipments / total_shipments_raw * 100.0) if total_shipments_raw else 0.0

c1, c2, c3 = st.columns(3)
c1.metric("Total Shipments (TL rows in file)",
          f"{total_shipments_raw:,}" if total_shipments_raw is not None else "N/A")
c2.metric("Eligible Shipments (after current rules)", f"{eligible_shipments:,}")
c3.metric("Coverage vs total TL shipments", f"{coverage:.1f}%")

if whole_journey:
    st.info(
        "Whole journey mode is ON. Each unique stop chain is treated as its own lane, and "
        "shipments must have every stop arrival + departure timestamp to qualify."
    )

if eligible_shipments == 0:
    diag = getattr(shipment_lt_all, "attrs", {}).get("diagnostics", {}) if shipment_lt_all is not None else {}
    msg_lines = [
        "No shipments are eligible for the current settings."
    ]
    if diag:
        rows_with_lane = diag.get("rows_with_lane", 0)
        rows_with_start = diag.get("rows_with_start_ts", 0)
        rows_with_end = diag.get("rows_with_end_ts", 0)
        sms = diag.get("start_milestone", start_ms)
        ems = diag.get("end_milestone", end_ms)
        msg_lines.append(
            f"Rows with a valid lane: {rows_with_lane:,}. "
            f"Rows with a usable start timestamp ({MILESTONE_LABELS.get(sms, sms)}): {rows_with_start:,}. "
            f"Rows with a usable end timestamp ({MILESTONE_LABELS.get(ems, ems)}): {rows_with_end:,}."
        )
        if rows_with_lane > 0 and (rows_with_start == 0 or rows_with_end == 0):
            msg_lines.append(
                "It looks like the chosen milestone timestamps are missing for every row. "
                "Try a different start/end milestone pair (e.g. Origin Arrival → Destination Arrival)."
            )
    msg_lines.append(
        "If everything looks correct, try toggling Whole Journey off, or pick a different start/end milestone pair."
    )
    st.warning(" ".join(msg_lines))

# ----------- Counts -----------
lane_counts, carrier_counts = compute_lane_and_carrier_counts(shipment_lt)

st.subheader("Lane & Carrier Counts (shipment volume)")
lc, cc = st.columns(2)
with lc:
    st.markdown("**Lane Counts**")
    st.caption(f"Unique lanes: {lane_counts.shape[0]:,}")
    st.dataframe(lane_counts.head(25), use_container_width=True)
    if lane_counts.shape[0] > 25:
        st.caption("Showing Top 25 lanes by shipment volume.")
with cc:
    st.markdown("**Carrier Counts**")
    st.caption(f"Unique carriers: {carrier_counts.shape[0]:,}")
    st.dataframe(carrier_counts.head(25), use_container_width=True)
    if carrier_counts.shape[0] > 25:
        st.caption("Showing Top 25 carriers by shipment volume.")

# Counts export
counts_key = build_key_glossary([], int(percentile_p))
counts_excel = write_excel_counts(lane_counts=lane_counts, carrier_counts=carrier_counts, key_df=counts_key)
counts_zip = write_csv_zip(
    ("lane_counts.csv", lane_counts),
    ("carrier_counts.csv", carrier_counts),
    key_df=counts_key,
)
ce1, ce2 = st.columns(2)
ce1.download_button(
    label="Download Lane + Carrier Counts (Excel)",
    data=counts_excel,
    file_name="lane_and_carrier_counts.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
ce2.download_button(
    label="Download Lane + Carrier Counts (CSV Zip)",
    data=counts_zip,
    file_name="lane_and_carrier_counts.zip",
    mime="application/zip",
)

# ----------- Shipment preview -----------
st.subheader("Shipment-level lead times (preview)")
preview_cols = [
    "TENANT_NAME", "MASTER_SHIPMENT_ID", "POL", "POD", "LANE",
    "CARRIER_NAME", "CARRIER_SCAC", "JOURNEY_LEAD_HOURS",
]
if whole_journey:
    seg_cols = [c for c in shipment_lt.columns if c.startswith("DWELL_") or c.startswith("SEG_")]
    preview_cols.extend(seg_cols)
preview_cols = [c for c in preview_cols if c in shipment_lt.columns]
st.dataframe(shipment_lt[preview_cols].head(25), use_container_width=True)
if shipment_lt.shape[0] > 25:
    st.caption(f"Showing 25 of {shipment_lt.shape[0]:,} shipment rows.")

# ----------- Final report -----------
duration_configs = build_duration_configs(start_ms, end_ms, whole_journey, shipment_lt)
report_df = build_carrier_lane_report(
    shipment_lt=shipment_lt,
    percentile_p=int(percentile_p),
    include_percentile=bool(include_percentile),
    min_volume_for_percentile=0,
    duration_configs=duration_configs,
)

st.subheader("Carrier Lane Lead (preview)")
preview_report = report_df.drop(columns=["_POL", "_POD", "_IS_LANE_ROW"], errors="ignore").head(25)
st.dataframe(preview_report, use_container_width=True)
if report_df.shape[0] > 25:
    st.caption(f"Showing 25 of {report_df.shape[0]:,} report rows. Full data available in the downloads below.")

# ----------- Final-report downloads (deferred) -----------
# The previous version built the full Excel + CSV ZIP on every page render
# even before the user clicked anything. On large uploads (60k+ TL rows)
# this was the OOM crash trigger because the openpyxl writer + raw_df copy
# easily consumed 700+ MB. Now the build only runs when the user clicks
# the prepare button, and the heavy "Raw Data" sheet is opt-in.

if "final_export_ready" not in st.session_state:
    st.session_state["final_export_ready"] = False
    st.session_state["final_export_bytes_xlsx"] = None
    st.session_state["final_export_bytes_zip"] = None
    st.session_state["final_export_key"] = None

st.markdown("### Final report downloads")
_RAW_DATA_WARN_ROWS = 25_000  # Rough threshold past which the openpyxl raw sheet write balloons memory.
include_raw_in_export = st.checkbox(
    "Include the full Raw Data sheet/file in the download (slower, bigger file)",
    value=False,
    help=(
        "When checked, the full input file is embedded in the Excel as a 'Raw Data' "
        "sheet and added to the CSV ZIP. Disable for large uploads to keep downloads "
        "small and fast."
    ),
)
if include_raw_in_export and len(raw_df) > _RAW_DATA_WARN_ROWS:
    st.warning(
        f"Heads up: this file has {len(raw_df):,} TL rows. Including the Raw Data sheet "
        f"can use 1.5-2x more memory while the Excel is being built and may exceed "
        f"Streamlit Cloud's memory limit. If the export fails, uncheck this option."
    )

# Use a key combining shape + settings so cached bytes invalidate when inputs change.
_export_key = (
    int(report_df.shape[0]),
    int(raw_df.shape[0]) if raw_df is not None else 0,
    str(start_ms), str(end_ms), bool(whole_journey),
    int(percentile_p), bool(include_percentile),
    int(top_n_lanes), bool(include_raw_in_export),
)
if st.session_state["final_export_key"] != _export_key:
    # Settings changed - clear any stale bytes so the user has to re-prepare.
    st.session_state["final_export_ready"] = False
    st.session_state["final_export_bytes_xlsx"] = None
    st.session_state["final_export_bytes_zip"] = None

if not st.session_state["final_export_ready"]:
    if st.button("Prepare final-report downloads", type="primary"):
        with st.spinner("Building Excel + CSV ZIP. This can take a minute on large uploads..."):
            key_df = build_key_glossary(duration_configs, int(percentile_p))
            xlsx_bytes = write_excel_final(
                raw_df=raw_df, report_df=report_df,
                duration_configs=duration_configs, percentile_p=int(percentile_p),
                key_df=key_df, include_raw_data=bool(include_raw_in_export),
            )
            display_report = report_df.drop(
                columns=["_IS_LANE_ROW", "_POL", "_POD"], errors="ignore"
            )
            zip_named = []
            if include_raw_in_export:
                zip_named.append(("raw_data.csv", raw_df))
            zip_named.append(("carrier_lane_lead.csv", display_report))
            zip_bytes = write_csv_zip(*zip_named, key_df=key_df)
            st.session_state["final_export_bytes_xlsx"] = xlsx_bytes
            st.session_state["final_export_bytes_zip"] = zip_bytes
            st.session_state["final_export_key"] = _export_key
            st.session_state["final_export_ready"] = True
        st.rerun()
    else:
        st.caption(
            "Click \"Prepare final-report downloads\" to build the Excel and CSV ZIP. "
            "Building is deferred so the page renders fast on large uploads."
        )

if st.session_state["final_export_ready"]:
    fe1, fe2 = st.columns(2)
    fe1.download_button(
        label="Download Final Excel Report",
        data=st.session_state["final_export_bytes_xlsx"],
        file_name=f"tl_carrier_lane_lead_{start_ms}_to_{end_ms}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fe2.download_button(
        label="Download Final Report (CSV Zip)",
        data=st.session_state["final_export_bytes_zip"],
        file_name=f"tl_carrier_lane_lead_{start_ms}_to_{end_ms}.zip",
        mime="application/zip",
    )

# ----------- Insights -----------
if "show_insights" not in st.session_state:
    st.session_state["show_insights"] = False

if st.button("Generate Insights"):
    st.session_state["show_insights"] = True

if st.session_state["show_insights"]:
    st.subheader("Insights")

    insight_options = build_insight_options(duration_configs)
    default_metric_label = list(insight_options.keys())[0]
    selected_metric_label = st.selectbox(
        "Choose journey part for insights",
        options=list(insight_options.keys()),
        index=list(insight_options.keys()).index(default_metric_label),
    )
    selected_metric_cfg = insight_options[selected_metric_label]

    with st.spinner("Computing insights..."):
        lane_summary_df, carrier_recs_df = compute_insights_for_metric(
            shipment_lt=shipment_lt,
            metric_cfg=selected_metric_cfg,
            percentile_p=int(percentile_p),
            percentile_threshold_enabled=bool(include_percentile and limit_by_volume),
            percentile_threshold_pct=float(percentile_volume_threshold_pct),
            rec_threshold_enabled=bool(recommendation_threshold_enabled),
            rec_threshold_pct=float(recommendation_threshold_pct),
        )

    lane_mapping = make_lane_selector_labels(lane_summary_df)

    if not lane_mapping:
        st.warning("No insight data available for the selected settings.")
    else:
        lane_labels = list(lane_mapping.keys())
        selected_lane_label = st.selectbox("Choose lane", options=lane_labels, index=0)
        selected_tenant, selected_lane = lane_mapping[selected_lane_label]

        lane_row, top5_df, selected_lane_shipments = get_selected_lane_outputs(
            shipment_lt=shipment_lt,
            lane_summary=lane_summary_df,
            carrier_recs=carrier_recs_df,
            selected_tenant=selected_tenant,
            selected_lane=selected_lane,
            metric_cfg=selected_metric_cfg,
        )

        if lane_row.empty:
            st.warning("No lane summary available for the selected lane.")
        else:
            lane_info = lane_row.iloc[0]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Lane Shipments", f"{int(lane_info['LANE_SHIPMENTS']):,}")
            c2.metric("Carriers in Lane", f"{int(lane_info['CARRIER_COUNT']):,}")
            c3.metric("Lane Median",
                      f"{lane_info['LANE_MEDIAN_D']} d" if pd.notna(lane_info["LANE_MEDIAN_D"]) else "N/A")
            c4.metric(f"Lane P{int(percentile_p)}",
                      f"{lane_info['LANE_PXX_D']} d" if pd.notna(lane_info["LANE_PXX_D"]) else "N/A")

            st.caption(
                "Recommendations are ranked by lowest deviation from the lane median, "
                "then lowest percentile deviation, then higher shipment volume. "
                "Thresholds (if enabled) are applied as carrier share % of total shipments in the selected lane."
            )

            display_top5 = top5_df[[
                "RANK_IN_LANE", "CARRIER_NAME", "CARRIER_SCAC", "SHIPMENTS",
                "CARRIER_SHARE_PCT", "CARRIER_MEDIAN_H", "CARRIER_MEDIAN_D",
                "CARRIER_PXX_H", "CARRIER_PXX_D",
                "MEDIAN_ABS_DEV_H", "MEDIAN_ABS_DEV_D",
                "PXX_ABS_DEV_H", "PXX_ABS_DEV_D",
            ]].copy()
            display_top5 = display_top5.rename(columns={
                "RANK_IN_LANE": "Rank",
                "CARRIER_NAME": "Carrier Name",
                "CARRIER_SCAC": "Carrier SCAC",
                "SHIPMENTS": "Shipments",
                "CARRIER_SHARE_PCT": "Carrier Share (%)",
                "CARRIER_MEDIAN_H": "Carrier Median (Hours)",
                "CARRIER_MEDIAN_D": "Carrier Median (Days)",
                "CARRIER_PXX_H": f"Carrier P{int(percentile_p)} (Hours)",
                "CARRIER_PXX_D": f"Carrier P{int(percentile_p)} (Days)",
                "MEDIAN_ABS_DEV_H": "Median Abs Deviation (Hours)",
                "MEDIAN_ABS_DEV_D": "Median Abs Deviation (Days)",
                "PXX_ABS_DEV_H": f"P{int(percentile_p)} Abs Deviation (Hours)",
                "PXX_ABS_DEV_D": f"P{int(percentile_p)} Abs Deviation (Days)",
            })
            st.markdown("**Top 5 Recommended Carriers**")
            st.dataframe(display_top5, use_container_width=True)

            if not top5_df.empty:
                bar_df = top5_df.copy()
                bar_df["Carrier Label"] = bar_df["CARRIER_NAME"].astype(str) + " (" + bar_df["CARRIER_SCAC"].fillna("").astype(str) + ")"
                fig_bar = px.bar(
                    bar_df.sort_values(["RANK_IN_LANE"]),
                    x="Carrier Label", y="MEDIAN_ABS_DEV_H",
                    hover_data={
                        "SHIPMENTS": True, "CARRIER_SHARE_PCT": True,
                        "CARRIER_MEDIAN_H": True, "CARRIER_PXX_H": True,
                        "PXX_ABS_DEV_H": True, "Carrier Label": False,
                        "MEDIAN_ABS_DEV_H": True,
                    },
                    labels={
                        "Carrier Label": "Carrier",
                        "MEDIAN_ABS_DEV_H": "Median Absolute Deviation (Hours)",
                    },
                    title=f"Deviation Ranking: {selected_metric_label} | {selected_lane}",
                )
                st.plotly_chart(fig_bar, use_container_width=True)

                if not selected_lane_shipments.empty:
                    ship_plot = selected_lane_shipments.copy()
                    carrier_order = top5_df["CARRIER_NAME"].tolist()
                    ship_plot["CARRIER_NAME"] = pd.Categorical(ship_plot["CARRIER_NAME"], categories=carrier_order, ordered=True)
                    ship_plot = ship_plot.sort_values("CARRIER_NAME")
                    lane_median_h = lane_info["LANE_MEDIAN_H"]
                    fig_box = px.box(
                        ship_plot, x="CARRIER_NAME", y="LEAD_TIME_HOURS",
                        points="outliers",
                        labels={
                            "CARRIER_NAME": "Carrier",
                            "LEAD_TIME_HOURS": f"{selected_metric_label} Lead Time (Hours)",
                        },
                        title=f"Lead Time Distribution by Carrier: {selected_metric_label} | {selected_lane}",
                    )
                    if pd.notna(lane_median_h):
                        fig_box.add_hline(
                            y=lane_median_h, line_dash="dash",
                            annotation_text=f"Lane Median: {lane_median_h} h",
                        )
                    st.plotly_chart(fig_box, use_container_width=True)

            insights_excel = write_insights_excel(
                lane_summary=lane_summary_df,
                carrier_recs=carrier_recs_df,
                selected_lane_shipments=selected_lane_shipments,
                key_df=key_df,
            )
            insights_zip = write_csv_zip(
                ("lane_summary.csv", lane_summary_df),
                ("carrier_recommendations.csv", carrier_recs_df),
                ("selected_lane_shipments.csv", selected_lane_shipments),
                key_df=key_df,
            )

            safe_metric_label = re.sub(r"[^A-Za-z0-9._-]", "_", selected_metric_label)
            ie1, ie2 = st.columns(2)
            ie1.download_button(
                label="Download Insights Excel",
                data=insights_excel,
                file_name=f"tl_insights_{safe_metric_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            ie2.download_button(
                label="Download Insights (CSV Zip)",
                data=insights_zip,
                file_name=f"tl_insights_{safe_metric_label}.zip",
                mime="application/zip",
            )
